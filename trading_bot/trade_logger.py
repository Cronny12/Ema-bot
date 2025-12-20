"""
Trade Logger Module
===================
Handles logging trades to Excel spreadsheet and console.
"""

import os
from datetime import datetime
from typing import Dict, List, Optional, Any
import logging
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

import config

logger = logging.getLogger(__name__)


class TradeLogger:
    """Logs trades to Excel spreadsheet."""
    
    TRADE_COLUMNS = [
        'Trade ID',
        'Timestamp',
        'Symbol',
        'Side',
        'Entry Price',
        'Qty',
        'Stop Price',
        'Risk $',
        'Risk %',
        'Exit Price',
        'Exit Time',
        'Exit Reason',
        'PnL $',
        'PnL %',
        'R Multiple',
        'Slippage (bps)',
        'Bars Held',
        'Signal Type',
        'EMA Fast',
        'EMA Slow',
        'RSI',
        'ADX',
        'ATR %',
        'Regime',
        'Notes'
    ]
    
    SUMMARY_COLUMNS = [
        'Date',
        'Starting Equity',
        'Ending Equity',
        'Daily PnL $',
        'Daily PnL %',
        'Trades',
        'Winners',
        'Losers',
        'Win Rate',
        'Avg Win $',
        'Avg Loss $',
        'Profit Factor',
        'Max Drawdown %',
        'Slippage Total',
        'Errors',
        'Equity Curve Slope'
    ]
    
    SHADOW_COLUMNS = [
        'Timestamp',
        'Symbol',
        'Signal Type',
        'Filtered By',
        'Entry Price',
        'Theoretical Stop',
        'Theoretical Size',
        'EMA Fast',
        'EMA Slow',
        'RSI',
        'ADX',
        'ATR %',
        'Notes'
    ]
    
    def __init__(self, filepath: str = None):
        self.filepath = filepath or os.path.join(
            config.LOG_DIR, config.TRADE_LOG_FILE
        )
        self.trade_counter = 0
        self._ensure_log_directory()
        self._initialize_workbook()
    
    def _ensure_log_directory(self):
        """Create log directory if it doesn't exist."""
        os.makedirs(config.LOG_DIR, exist_ok=True)
    
    def _initialize_workbook(self):
        """Initialize or load the Excel workbook."""
        if os.path.exists(self.filepath):
            try:
                self.wb = load_workbook(self.filepath)
                # Get last trade ID
                if 'Trades' in self.wb.sheetnames:
                    ws = self.wb['Trades']
                    if ws.max_row > 1:
                        last_id = ws.cell(row=ws.max_row, column=1).value
                        if last_id and isinstance(last_id, str) and last_id.startswith('T'):
                            self.trade_counter = int(last_id[1:])
            except Exception as e:
                logger.warning(f"Could not load existing workbook: {e}")
                self._create_new_workbook()
        else:
            self._create_new_workbook()
    
    def _create_new_workbook(self):
        """Create a new workbook with all required sheets."""
        self.wb = Workbook()
        
        # Trades sheet
        ws_trades = self.wb.active
        ws_trades.title = 'Trades'
        self._setup_sheet(ws_trades, self.TRADE_COLUMNS)
        
        # Daily Summary sheet
        ws_summary = self.wb.create_sheet('Daily Summary')
        self._setup_sheet(ws_summary, self.SUMMARY_COLUMNS)
        
        # Shadow Mode sheet (filtered signals)
        ws_shadow = self.wb.create_sheet('Shadow Mode')
        self._setup_sheet(ws_shadow, self.SHADOW_COLUMNS)
        
        # Alternative Exits sheet
        ws_alts = self.wb.create_sheet('Alternative Exits')
        alt_columns = ['Trade ID', 'Exit Strategy', 'Theoretical Exit Price', 
                       'Theoretical PnL $', 'Theoretical R Multiple', 'Notes']
        self._setup_sheet(ws_alts, alt_columns)
        
        self._save()
    
    def _setup_sheet(self, ws, columns: List[str]):
        """Setup a sheet with headers and formatting."""
        # Header styling
        header_font = Font(bold=True, color='FFFFFF')
        header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
        header_alignment = Alignment(horizontal='center', vertical='center')
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Write headers
        for col, header in enumerate(columns, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border
        
        # Set column widths
        for col in range(1, len(columns) + 1):
            ws.column_dimensions[get_column_letter(col)].width = 15
        
        # Freeze header row
        ws.freeze_panes = 'A2'
    
    def _save(self):
        """Save workbook to file."""
        try:
            self.wb.save(self.filepath)
        except Exception as e:
            logger.error(f"Error saving trade log: {e}")
    
    def _generate_trade_id(self) -> str:
        """Generate unique trade ID."""
        self.trade_counter += 1
        return f"T{self.trade_counter:05d}"
    
    def log_entry(
        self,
        symbol: str,
        side: str,
        entry_price: float,
        qty: int,
        stop_price: float,
        risk_amount: float,
        risk_percent: float,
        signal_type: str,
        ema_fast: int,
        ema_slow: int,
        rsi: float,
        adx: float,
        atr_percent: float,
        regime: str,
        notes: str = ""
    ) -> str:
        """
        Log a trade entry.
        Returns trade ID.
        """
        trade_id = self._generate_trade_id()
        timestamp = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        
        ws = self.wb['Trades']
        row = ws.max_row + 1
        
        # Write entry data
        data = [
            trade_id,
            timestamp,
            symbol,
            side,
            entry_price,
            qty,
            stop_price,
            risk_amount,
            f"{risk_percent:.2%}",
            None,  # Exit Price (filled on exit)
            None,  # Exit Time
            None,  # Exit Reason
            None,  # PnL $
            None,  # PnL %
            None,  # R Multiple
            None,  # Slippage
            None,  # Bars Held
            signal_type,
            ema_fast,
            ema_slow,
            f"{rsi:.1f}",
            f"{adx:.1f}",
            f"{atr_percent:.4f}",
            regime,
            notes
        ]
        
        for col, value in enumerate(data, 1):
            ws.cell(row=row, column=col, value=value)
        
        self._save()
        
        logger.info(f"Trade entry logged: {trade_id} {side} {qty} {symbol} @ ${entry_price:.2f}")
        
        return trade_id
    
    def log_exit(
        self,
        trade_id: str,
        exit_price: float,
        exit_reason: str,
        pnl_dollars: float,
        pnl_percent: float,
        r_multiple: float,
        slippage_bps: float,
        bars_held: int
    ):
        """Log a trade exit."""
        ws = self.wb['Trades']
        
        # Find the trade row
        trade_row = None
        for row in range(2, ws.max_row + 1):
            if ws.cell(row=row, column=1).value == trade_id:
                trade_row = row
                break
        
        if trade_row is None:
            logger.error(f"Trade ID {trade_id} not found for exit logging")
            return
        
        exit_time = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        
        # Update exit columns
        ws.cell(row=trade_row, column=10, value=exit_price)
        ws.cell(row=trade_row, column=11, value=exit_time)
        ws.cell(row=trade_row, column=12, value=exit_reason)
        ws.cell(row=trade_row, column=13, value=pnl_dollars)
        ws.cell(row=trade_row, column=14, value=f"{pnl_percent:.2%}")
        ws.cell(row=trade_row, column=15, value=f"{r_multiple:.2f}R")
        ws.cell(row=trade_row, column=16, value=f"{slippage_bps:.1f}")
        ws.cell(row=trade_row, column=17, value=bars_held)
        
        # Color code P&L
        pnl_cell = ws.cell(row=trade_row, column=13)
        if pnl_dollars >= 0:
            pnl_cell.fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
        else:
            pnl_cell.fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
        
        self._save()
        
        logger.info(f"Trade exit logged: {trade_id} @ ${exit_price:.2f} | PnL: ${pnl_dollars:.2f} ({r_multiple:.2f}R)")
    
    def log_shadow_signal(
        self,
        symbol: str,
        signal_type: str,
        filtered_by: str,
        entry_price: float,
        theoretical_stop: float,
        theoretical_size: int,
        ema_fast: int,
        ema_slow: int,
        rsi: float,
        adx: float,
        atr_percent: float,
        notes: str = ""
    ):
        """Log a filtered signal for shadow mode analysis."""
        ws = self.wb['Shadow Mode']
        row = ws.max_row + 1
        
        timestamp = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        
        data = [
            timestamp,
            symbol,
            signal_type,
            filtered_by,
            entry_price,
            theoretical_stop,
            theoretical_size,
            ema_fast,
            ema_slow,
            f"{rsi:.1f}",
            f"{adx:.1f}",
            f"{atr_percent:.4f}",
            notes
        ]
        
        for col, value in enumerate(data, 1):
            ws.cell(row=row, column=col, value=value)
        
        self._save()
    
    def log_alternative_exit(
        self,
        trade_id: str,
        exit_strategy: str,
        theoretical_exit_price: float,
        theoretical_pnl: float,
        theoretical_r: float,
        notes: str = ""
    ):
        """Log alternative exit strategy results for comparison."""
        ws = self.wb['Alternative Exits']
        row = ws.max_row + 1
        
        data = [
            trade_id,
            exit_strategy,
            theoretical_exit_price,
            theoretical_pnl,
            f"{theoretical_r:.2f}R",
            notes
        ]
        
        for col, value in enumerate(data, 1):
            ws.cell(row=row, column=col, value=value)
        
        self._save()
    
    def log_daily_summary(
        self,
        date_str: str,
        starting_equity: float,
        ending_equity: float,
        daily_pnl: float,
        daily_pnl_percent: float,
        trades: int,
        winners: int,
        losers: int,
        avg_win: float,
        avg_loss: float,
        max_drawdown: float,
        total_slippage: float,
        errors: int,
        equity_slope: float
    ):
        """Log end-of-day summary."""
        ws = self.wb['Daily Summary']
        row = ws.max_row + 1
        
        win_rate = winners / trades if trades > 0 else 0
        profit_factor = (avg_win * winners) / (abs(avg_loss) * losers) if losers > 0 and avg_loss != 0 else 0
        
        data = [
            date_str,
            starting_equity,
            ending_equity,
            daily_pnl,
            f"{daily_pnl_percent:.2%}",
            trades,
            winners,
            losers,
            f"{win_rate:.1%}",
            avg_win,
            avg_loss,
            f"{profit_factor:.2f}",
            f"{max_drawdown:.2%}",
            total_slippage,
            errors,
            f"{equity_slope:.4f}"
        ]
        
        for col, value in enumerate(data, 1):
            ws.cell(row=row, column=col, value=value)
        
        # Color code daily P&L
        pnl_cell = ws.cell(row=row, column=4)
        if daily_pnl >= 0:
            pnl_cell.fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
        else:
            pnl_cell.fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
        
        self._save()
        
        logger.info(f"Daily summary logged for {date_str}: PnL ${daily_pnl:.2f}, {trades} trades, {win_rate:.1%} win rate")
    
    def get_trade_history(self, days: int = 30) -> List[Dict[str, Any]]:
        """Get recent trade history."""
        ws = self.wb['Trades']
        trades = []
        
        for row in range(2, ws.max_row + 1):
            trade = {}
            for col, header in enumerate(self.TRADE_COLUMNS, 1):
                trade[header] = ws.cell(row=row, column=col).value
            trades.append(trade)
        
        return trades[-days*10:] if len(trades) > days*10 else trades  # Approximate
    
    def get_statistics(self) -> Dict[str, Any]:
        """Calculate overall trading statistics."""
        ws = self.wb['Trades']
        
        total_trades = 0
        winners = 0
        losers = 0
        total_pnl = 0.0
        total_win_pnl = 0.0
        total_loss_pnl = 0.0
        
        for row in range(2, ws.max_row + 1):
            pnl = ws.cell(row=row, column=13).value
            if pnl is not None:
                total_trades += 1
                total_pnl += pnl
                if pnl >= 0:
                    winners += 1
                    total_win_pnl += pnl
                else:
                    losers += 1
                    total_loss_pnl += pnl
        
        return {
            'total_trades': total_trades,
            'winners': winners,
            'losers': losers,
            'win_rate': winners / total_trades if total_trades > 0 else 0,
            'total_pnl': total_pnl,
            'avg_win': total_win_pnl / winners if winners > 0 else 0,
            'avg_loss': total_loss_pnl / losers if losers > 0 else 0,
            'profit_factor': total_win_pnl / abs(total_loss_pnl) if total_loss_pnl != 0 else 0
        }
